import os
import json
import nltk
from nltk import word_tokenize, sent_tokenize
import openpyxl
from derived_variables_extractor import *

cwd = os.getcwd()

modified_files_path = os.path.join(cwd, 'clean_files')
mod_files = os.listdir(modified_files_path)

for file in mod_files:
    
    filepath = os.path.join(modified_files_path, file)

    print('-\n' + f'\nAnalyzing file "{file}"' + '\n')
    
    with open(filepath, 'r+', encoding='utf-8') as currentfile:
        data = currentfile.read()
        # tokens = words in txt
        tokens = word_tokenize(data)
        excluded_values = [',', '.', ':', '%', '(', ')', '"', '“', '”', '’', '?', '!']
        for val in excluded_values: 
            while val in tokens:
                tokens.remove(val)
        sentences = sent_tokenize(data)
    
        print('-')
        
        posScore = calculatePositiveScore(tokens)
        print('Positive Score: ', posScore)
        negScore = calculateNegativeScore(tokens)            
        print('Negative Score: ', negScore)
            
        polScore = calculatePolarityScore(tokens)
        subScore = calculateSubjectivityScore(tokens)
        avgSentenceLen = calculateAverageSentenceLength(tokens, sentences)
        json_string = calculateComplexWords(tokens)
        complex_obj = json.loads(json_string)
        fogIndex = calculateFogIndex(tokens, sentences)
        avgWordPerSentence = calculateAverageWordPerSentence(tokens, sentences)
        syllablesPerWord = count_syllables(tokens)
        tokens_string = ', '.join(tokens)
        personalPronouns = calculatePersonalPronouns(tokens_string)
        avgWordLength = calculateAverageWordLength(tokens)

        print('-\n')

        workbook = openpyxl.load_workbook(os.path.join(cwd, 'output', 'Output Data Structure.xlsx'))
        worksheet = workbook.active
        
        for row in worksheet.iter_rows(min_row = 2, max_row = 115, min_col = 2, max_col = 2):
            for url_cell in row:
                url = url_cell.value
                name = url.replace('https://insights.blackcoffer.com/', '').replace('/', '')
                if name.replace('.txt', '') == file.replace('.txt', ''):
                    for column in worksheet.iter_cols(min_row = url_cell.row, max_row = url_cell.row, min_col = 3, max_col = 15):
                        for cell in column:
                            cellLetter = cell.coordinate[0]
                            if cellLetter == 'C':
                                cell.value = posScore
                            elif cellLetter == 'D':
                                cell.value = negScore
                            elif cellLetter == 'E':
                                cell.value = polScore
                            elif cellLetter == 'F':
                                cell.value = subScore
                            elif cellLetter == 'G':
                                cell.value = avgWordPerSentence
                            elif cellLetter == 'H':
                                cell.value = complex_obj['percentage']
                            elif cellLetter == 'I':
                                cell.value = fogIndex
                            elif cellLetter == 'J':
                                cell.value = avgWordPerSentence
                            elif cellLetter == 'K':
                                cell.value = complex_obj['count']
                            elif cellLetter == 'L':
                                cell.value = len(tokens)
                            elif cellLetter == 'M':
                                cell.value = syllablesPerWord
                            elif cellLetter == 'N':
                                cell.value = personalPronouns
                            elif cellLetter == 'O':
                                cell.value = avgWordLength

        output_filepath = os.path.join(cwd, 'output', 'Output Data Structure.xlsx')
        workbook.save(output_filepath)